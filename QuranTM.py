import openpyxl
import pprint
import re

wb = openpyxl.load_workbook('QuranTM.xlsm')
Quran_db = wb['QuranTM']
row = Quran_db.max_row
column = 6
dslist = []
dslistfinal = []
intlist = []
for r in range(3, row + 1):
    for c in range(1, column + 1):
        ds = Quran_db.cell(r, c).value
        if ds is not None:
            found = ""
            m = re.search('Root(.*) ]', ds)
        if m:
            found = m.group(1)
            ds = found
            ds = re.sub(" *", "", ds)
            dslist.append(ds)
for ds in dslist:
    if ds not in dslistfinal:
        dslistfinal.append(ds)
for i in range(1, 1672):
    intlist.append(i)
dictdslistfinal = dict(zip(intlist, dslistfinal))
print('Here is the list of roots for which Allah SWT chose words in Quran')
print('--- is the entry for Prepositions, Conjunctions, Emphasis particles, etc. which have no roots\n')
pprint.pprint(dictdslistfinal)
# print(dictdslistfinal)
print('\n')
dr = ""


def display():
    for r in range(3, row + 1):
        for c in range(1, column + 1):
            ds = Quran_db.cell(r, c).value
            if ds is not None:
                found = ""
                m = re.search('Root(.*) ]', ds)
            if m:
                found = m.group(1)
                ds = found
                ds = re.sub(" *", "", ds)
                # if ds == dr:
                if ds == dictdslistfinal[int(dr)]:
                    dw = Quran_db.cell(r - 1, c).value
                    dw1 = Quran_db.cell(r - 1, 1).value
                    dw2 = Quran_db.cell(r - 1, 2).value
                    dw3 = Quran_db.cell(r - 1, 3).value
                    dw4 = Quran_db.cell(r - 1, 4).value
                    dw5 = Quran_db.cell(r - 1, 5).value
                    dw6 = Quran_db.cell(r - 1, 6).value
                    ds1 = Quran_db.cell(r, c).value
                    ds1 = re.sub('\[.*\]', "", ds1)
                    nrc = "?,-()._x000D_\""
                    for char in nrc:
                        ds1 = ds1.replace(char, "")  # english phrase left
                    print('\n', dw, '\n', ds1, dw1, dw2, dw3, dw4, dw5, dw6)


choice = 'y'
search = input("\nWould you like to explore the Allah\'s wording and their meanings in Quran? " + "\nEnter y for Yes or any other key to quit the program : ")
while search.lower() == 'y':
    if choice == 'y':
        dr = input("Enter the root number from the list above ")
        try:
            if int(dr) in dictdslistfinal.keys():
                display()
            else:
                print("Sorry, the number must be amongst the list above")
        except:
            print("Ooops! integer expected")
    choice = input("\nYou may enter y again to choose another root number or enter any other key to exit the program : ")
    if choice.lower() == 'y':
        search = choice
    else:
        break
